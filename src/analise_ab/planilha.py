"""Geracao da planilha executiva recriada a cada analise."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from datetime import date
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet


COR_VERDE = "00856A"
COR_VERDE_CLARO = "DDF3EC"
COR_AMARELO = "F4B400"
COR_AMARELO_CLARO = "FFF4CC"
COR_VERMELHO = "C43D4E"
COR_VERMELHO_CLARO = "FBE5E8"
COR_TEXTO = "263238"
COR_CINZA = "607D8B"
COR_CINZA_CLARO = "EEF2F4"
COR_BRANCO = "FFFFFF"
BORDA_CLARA = Side(style="thin", color="D5DEE2")

ROTULOS = {
    "teste_id": "Teste",
    "parceiro": "Parceiro",
    "data": "Data",
    "grupo": "Grupo",
    "fase_id": "Fase",
    "data_inicio": "Inicio",
    "data_fim": "Fim",
    "dias": "Dias",
    "quantidade_grupos": "Grupos",
    "quantidade_fases": "Fases",
    "taxa_comissao": "Taxa de comissao",
    "taxa_cashback": "Taxa de cashback",
    "taxa_margem": "Taxa de margem",
    "contraste_cashback": "Contraste de cashback",
    "configuracao_cashback": "Cashback por grupo",
    "status_fase": "Status da fase",
    "compradores": "Compradores",
    "compradores_dia": "Compradores por dia",
    "comissao": "Comissao (R$)",
    "cashback": "Cashback (R$)",
    "vendas_totais": "GMV (R$)",
    "vendas_dia": "GMV por dia (R$)",
    "margem": "Resultado direto (R$)",
    "margem_dia": "Resultado direto por dia (R$)",
    "ticket_medio": "Ticket medio (R$)",
    "margem_por_comprador": "Resultado por comprador (R$)",
    "projecao_margem_100_trafego": "Projecao para 100% (R$)",
    "dias_projecao": "Dias da projecao",
    "grupo_controle": "Controle",
    "grupo_variante": "Variante",
    "metrica": "Metrica",
    "dias_pareados": "Dias pareados",
    "media_controle": "Media do controle",
    "media_variante": "Media da variante",
    "diferenca_media": "Diferenca media",
    "uplift_percentual": "Uplift",
    "limite_inferior_95": "Limite inferior 95%",
    "limite_superior_95": "Limite superior 95%",
    "valor_p_permutacao": "Valor-p",
    "delta_compradores_dia": "Delta compradores por dia",
    "delta_vendas_dia": "Delta GMV por dia (R$)",
    "delta_cashback_dia": "Delta cashback por dia (R$)",
    "delta_margem_dia": "Delta resultado por dia (R$)",
    "custo_cashback_por_comprador_incremental": "Custo por comprador incremental (R$)",
    "uplift_gmv_observado": "Uplift de GMV observado",
    "uplift_gmv_break_even": "Uplift de GMV para break-even",
    "descricao": "Descricao",
    "resultado": "Resultado",
    "status": "Status",
    "variante_recomendada": "Variante recomendada",
    "decisao": "Decisao",
    "alertas": "Alertas",
    "periodo": "Periodo analisado",
    "configuracao": "Configuracao",
    "severidade": "Severidade",
    "tipo": "Tipo",
    "id_analise": "ID",
    "data_analise": "Data da analise",
    "nome_teste": "Nome do teste",
    "controle": "Controle",
    "variantes": "Variantes avaliadas",
    "metrica_principal": "Metrica principal",
    "resultado_direto": "Resultado direto (R$)",
    "status_qualidade": "Qualidade do teste",
    "caminho_relatorio": "Relatorio",
    "hash_dados": "Hash dos dados",
    "versao": "Versao",
    "leitura": "Leitura",
    "metodo": "Metodo",
    "limite_inferior_percentual": "Limite inferior 95%",
    "limite_superior_percentual": "Limite superior 95%",
    "uplift_compradores": "Uplift compradores",
    "uplift_gmv": "Uplift GMV",
    "uplift_resultado": "Uplift resultado direto",
    "conclusao_agente": "Conclusao do agente",
}


def _valor_excel(valor: object) -> object:
    """Converte escalares do pandas e numpy para tipos aceitos pelo Excel."""

    if valor is None or pd.isna(valor):
        return None
    if isinstance(valor, pd.Timestamp):
        return valor.to_pydatetime()
    if isinstance(valor, np.generic):
        return valor.item()
    return valor


def _nome_aba(nome: str, existentes: set[str]) -> str:
    """Cria um nome de aba valido e unico."""

    base = re.sub(r"[\\/*?:\[\]]", "-", nome).strip() or "Teste"
    base = base[:31]
    candidato = base
    contador = 2
    while candidato.lower() in existentes:
        sufixo = f" ({contador})"
        candidato = f"{base[:31 - len(sufixo)]}{sufixo}"
        contador += 1
    existentes.add(candidato.lower())
    return candidato


def _nome_tabela(nome: str, usados: set[str]) -> str:
    """Cria um identificador unico aceito por tabelas do Excel."""

    base = re.sub(r"[^A-Za-z0-9_]", "_", nome)
    if not base or base[0].isdigit():
        base = f"Tabela_{base}"
    candidato = base[:240]
    contador = 2
    while candidato.lower() in usados:
        sufixo = f"_{contador}"
        candidato = f"{base[:240 - len(sufixo)]}{sufixo}"
        contador += 1
    usados.add(candidato.lower())
    return candidato


def _formatacao_numero(coluna: str) -> str | None:
    """Escolhe a formatacao numerica de acordo com a metrica."""

    if coluna in {"data", "data_inicio", "data_fim", "data_analise"}:
        return "dd/mm/yyyy"
    if coluna.startswith("taxa_") or coluna in {
        "contraste_cashback",
        "uplift_percentual",
        "uplift_gmv_observado",
        "uplift_gmv_break_even",
        "limite_inferior_percentual",
        "limite_superior_percentual",
        "uplift_compradores",
        "uplift_gmv",
        "uplift_resultado",
    }:
        return "0.0%"
    if coluna in {
        "comissao",
        "cashback",
        "vendas_totais",
        "vendas_dia",
        "margem",
        "margem_dia",
        "ticket_medio",
        "margem_por_comprador",
        "projecao_margem_100_trafego",
        "delta_vendas_dia",
        "delta_cashback_dia",
        "delta_margem_dia",
        "custo_cashback_por_comprador_incremental",
        "resultado_direto",
    }:
        return '"R$" #,##0.00'
    if coluna in {
        "fase_id",
        "dias",
        "quantidade_grupos",
        "quantidade_fases",
        "compradores",
        "dias_pareados",
        "dias_projecao",
    }:
        return "#,##0"
    if coluna == "valor_p_permutacao":
        return "0.0000"
    return None


def _titulo(aba: Worksheet, texto: str, subtitulo: str, ultima_coluna: int = 8) -> None:
    """Cria o cabecalho visual de uma aba."""

    aba.merge_cells(start_row=1, start_column=1, end_row=2, end_column=ultima_coluna)
    celula = aba.cell(1, 1, texto)
    celula.font = Font(name="Aptos Display", size=20, bold=True, color=COR_BRANCO)
    celula.fill = PatternFill("solid", fgColor=COR_VERDE)
    celula.alignment = Alignment(vertical="center")
    aba.row_dimensions[1].height = 28
    aba.row_dimensions[2].height = 10

    aba.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ultima_coluna)
    celula_subtitulo = aba.cell(3, 1, subtitulo)
    celula_subtitulo.font = Font(size=10, color=COR_CINZA)
    celula_subtitulo.alignment = Alignment(vertical="center")
    aba.row_dimensions[3].height = 22
    aba.sheet_view.showGridLines = False


def _secao(aba: Worksheet, linha: int, texto: str, ultima_coluna: int) -> None:
    """Destaca o inicio de uma secao."""

    aba.merge_cells(
        start_row=linha,
        start_column=1,
        end_row=linha,
        end_column=max(1, ultima_coluna),
    )
    celula = aba.cell(linha, 1, texto)
    celula.font = Font(bold=True, color=COR_TEXTO, size=11)
    celula.fill = PatternFill("solid", fgColor=COR_CINZA_CLARO)
    celula.alignment = Alignment(vertical="center")
    aba.row_dimensions[linha].height = 22


def _escrever_tabela(
    aba: Worksheet,
    dados: pd.DataFrame,
    linha: int,
    coluna: int,
    nome: str,
    tabelas_usadas: set[str],
) -> tuple[int, int]:
    """Escreve um DataFrame como tabela formatada e devolve seus limites."""

    if dados.empty:
        aba.cell(linha, coluna, "Sem dados aplicaveis.")
        aba.cell(linha, coluna).font = Font(italic=True, color=COR_CINZA)
        return linha, coluna

    colunas = list(dados.columns)
    for deslocamento, nome_coluna in enumerate(colunas):
        celula = aba.cell(linha, coluna + deslocamento, ROTULOS.get(nome_coluna, nome_coluna))
        celula.font = Font(bold=True, color=COR_BRANCO)
        celula.fill = PatternFill("solid", fgColor=COR_VERDE)
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celula.border = Border(bottom=BORDA_CLARA)

    registros = dados.itertuples(index=False, name=None)
    for deslocamento_linha, registro in enumerate(registros, start=1):
        for deslocamento_coluna, valor in enumerate(registro):
            nome_coluna = colunas[deslocamento_coluna]
            celula = aba.cell(
                linha + deslocamento_linha,
                coluna + deslocamento_coluna,
                _valor_excel(valor),
            )
            celula.alignment = Alignment(vertical="top", wrap_text=True)
            formato = _formatacao_numero(nome_coluna)
            if formato is None and pd.api.types.is_numeric_dtype(dados[nome_coluna]):
                formato = "#,##0.00"
            if formato:
                celula.number_format = formato

    ultima_linha = linha + len(dados)
    ultima_coluna = coluna + len(colunas) - 1
    referencia = (
        f"{get_column_letter(coluna)}{linha}:"
        f"{get_column_letter(ultima_coluna)}{ultima_linha}"
    )
    tabela = Table(displayName=_nome_tabela(nome, tabelas_usadas), ref=referencia)
    tabela.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    aba.add_table(tabela)
    return ultima_linha, ultima_coluna


def _ajustar_colunas(aba: Worksheet, limite: int = 45) -> None:
    """Dimensiona as colunas sem deixar textos longos dominar a tela."""

    for indice in range(1, aba.max_column + 1):
        valores = [
            str(aba.cell(linha, indice).value)
            for linha in range(1, min(aba.max_row, 150) + 1)
            if aba.cell(linha, indice).value is not None
        ]
        largura = min(limite, max(11, max((len(valor) for valor in valores), default=8) + 2))
        aba.column_dimensions[get_column_letter(indice)].width = largura


def _grafico_barras(
    aba: Worksheet,
    linha_inicio: int,
    linha_fim: int,
    coluna_categorias: int,
    coluna_valores: int,
    titulo: str,
    ancora: str,
) -> None:
    """Adiciona um grafico de barras para uma tabela existente."""

    if linha_fim <= linha_inicio:
        return
    grafico = BarChart()
    grafico.type = "bar"
    grafico.style = 10
    grafico.title = titulo
    grafico.y_axis.title = "Configuracao"
    grafico.x_axis.title = "R$"
    grafico.x_axis.numFmt = '"R$" #,##0'
    grafico.height = min(12, max(7, (linha_fim - linha_inicio) * 0.55))
    grafico.width = 15
    valores = Reference(
        aba,
        min_col=coluna_valores,
        min_row=linha_inicio,
        max_row=linha_fim,
    )
    categorias = Reference(
        aba,
        min_col=coluna_categorias,
        min_row=linha_inicio + 1,
        max_row=linha_fim,
    )
    grafico.add_data(valores, titles_from_data=True)
    grafico.set_categories(categorias)
    grafico.legend = None
    aba.add_chart(grafico, ancora)


def _dados_decisoes(
    decisoes: pd.DataFrame,
    fases: pd.DataFrame,
    resumo: pd.DataFrame,
) -> pd.DataFrame:
    """Monta a tabela executiva com periodo, resultado e qualidade."""

    periodos = fases.groupby(["teste_id", "parceiro"], as_index=False).agg(
        data_inicio=("data_inicio", "min"),
        data_fim=("data_fim", "max"),
    )
    grupos = resumo.groupby(["teste_id", "parceiro"], as_index=False).agg(
        quantidade_grupos=("grupo", "nunique")
    )
    tabela = decisoes.merge(periodos, on=["teste_id", "parceiro"], how="left")
    tabela = tabela.merge(grupos, on=["teste_id", "parceiro"], how="left")
    principal = _resumo_fases_principais(resumo, decisoes)
    recomendadas = principal.merge(
        decisoes[["teste_id", "parceiro", "variante_recomendada"]],
        on=["teste_id", "parceiro"],
        how="left",
    )
    recomendadas = recomendadas[
        recomendadas["grupo"] == recomendadas["variante_recomendada"]
    ][["teste_id", "parceiro", "taxa_cashback", "margem", "taxa_margem"]]
    tabela = tabela.merge(recomendadas, on=["teste_id", "parceiro"], how="left")
    tabela["periodo"] = tabela.apply(
        lambda linha: (
            f"{linha['data_inicio']:%d/%m/%Y} a {linha['data_fim']:%d/%m/%Y}"
            if pd.notna(linha["data_inicio"]) and pd.notna(linha["data_fim"])
            else "-"
        ),
        axis=1,
    )
    tabela["status_qualidade"] = tabela.apply(_status_qualidade, axis=1)
    return tabela[
        [
            "teste_id",
            "parceiro",
            "periodo",
            "variante_recomendada",
            "taxa_cashback",
            "margem",
            "taxa_margem",
            "status",
            "decisao",
            "status_qualidade",
        ]
    ]


def _resumo_fases_principais(resumo: pd.DataFrame, decisoes: pd.DataFrame) -> pd.DataFrame:
    """Seleciona as variantes da fase usada pela regra de decisao."""

    chaves = decisoes[["teste_id", "parceiro", "fase_principal"]].rename(
        columns={"fase_principal": "fase_id"}
    )
    principal = resumo.merge(chaves, on=["teste_id", "parceiro", "fase_id"], how="inner")
    principal = principal.copy()
    principal.insert(
        0,
        "configuracao",
        principal["parceiro"].astype(str) + " - " + principal["grupo"].astype(str),
    )
    return principal


def _status_qualidade(decisao: pd.Series) -> str:
    """Resume os principais riscos de qualidade sem substituir os alertas."""

    if int(decisao.get("quantidade_fases", 1)) > 1:
        return "Atencao: tratamentos mudaram"
    if "expostos" in str(decisao.get("alertas", "")).lower():
        return "Atencao: validar exposicao"
    return "Dados consistentes"


def _leitura_intervalo(limite_inferior: float, limite_superior: float) -> str:
    """Classifica o intervalo de confianca pelo sinal do efeito."""

    if limite_inferior > 0:
        return "Ganho consistente"
    if limite_superior < 0:
        return "Queda consistente"
    return "Inconclusivo"


def _comparacoes_executivas(comparacoes: pd.DataFrame) -> pd.DataFrame:
    """Prepara os efeitos estatisticos em uma escala percentual comparavel."""

    if comparacoes.empty:
        return comparacoes.copy()
    tabela = comparacoes.copy()
    denominador = tabela["media_controle"].replace(0, np.nan)
    tabela["limite_inferior_percentual"] = tabela["limite_inferior_95"] / denominador
    tabela["limite_superior_percentual"] = tabela["limite_superior_95"] / denominador
    tabela["leitura"] = tabela.apply(
        lambda linha: _leitura_intervalo(
            float(linha["limite_inferior_95"]),
            float(linha["limite_superior_95"]),
        ),
        axis=1,
    )
    tabela["metodo"] = "Bootstrap diario pareado"
    return tabela[
        [
            "fase_id",
            "grupo_controle",
            "grupo_variante",
            "metrica",
            "dias_pareados",
            "uplift_percentual",
            "limite_inferior_percentual",
            "limite_superior_percentual",
            "valor_p_permutacao",
            "leitura",
            "metodo",
        ]
    ]


def _resumo_executivo_teste(
    resumo: pd.DataFrame,
    comparacoes: pd.DataFrame,
    decisao: pd.Series,
) -> pd.DataFrame:
    """Monta a tabela compacta da fase principal com totais e uplifts."""

    fase_principal = int(decisao["fase_principal"])
    tabela = resumo[resumo["fase_id"] == fase_principal].copy()
    colunas = [
        "grupo",
        "taxa_cashback",
        "compradores",
        "vendas_totais",
        "comissao",
        "cashback",
        "margem",
        "taxa_margem",
        "ticket_medio",
    ]
    tabela = tabela[colunas]

    efeitos = comparacoes[comparacoes["fase_id"] == fase_principal].pivot_table(
        index="grupo_variante",
        columns="metrica",
        values="uplift_percentual",
        aggfunc="first",
    )
    tabela = tabela.merge(efeitos, left_on="grupo", right_index=True, how="left")
    tabela = tabela.rename(
        columns={
            "compradores_x": "compradores",
            "compradores_y": "uplift_compradores",
            "vendas_totais_x": "vendas_totais",
            "vendas_totais_y": "uplift_gmv",
            "margem_y": "uplift_resultado",
            "margem_x": "margem",
        }
    )
    for coluna in ["uplift_compradores", "uplift_gmv", "uplift_resultado"]:
        if coluna not in tabela:
            tabela[coluna] = np.nan
    controle = str(comparacoes["grupo_controle"].iloc[0]) if not comparacoes.empty else "Grupo 1"
    tabela.loc[tabela["grupo"] == controle, [
        "uplift_compradores",
        "uplift_gmv",
        "uplift_resultado",
    ]] = 0.0

    recomendada = str(decisao["variante_recomendada"])
    tabela["decisao"] = "Nao escalar"
    tabela.loc[tabela["grupo"] == controle, "decisao"] = "Controle"
    if decisao["status"] == "escalar":
        tabela.loc[tabela["grupo"] == recomendada, "decisao"] = "Recomendada"
    return tabela[
        [
            "grupo",
            "taxa_cashback",
            "compradores",
            "vendas_totais",
            "comissao",
            "cashback",
            "margem",
            "taxa_margem",
            "ticket_medio",
            "uplift_compradores",
            "uplift_gmv",
            "uplift_resultado",
            "decisao",
        ]
    ]


def _acompanhamento_executivo(
    decisoes: pd.DataFrame,
    fases: pd.DataFrame,
    resumo: pd.DataFrame,
    dados: pd.DataFrame,
) -> pd.DataFrame:
    """Cria uma trilha de auditoria detalhada para cada teste analisado."""

    registros: list[dict[str, object]] = []
    for indice, decisao in decisoes.reset_index(drop=True).iterrows():
        teste_id = str(decisao["teste_id"])
        parceiro = str(decisao["parceiro"])
        filtro = (fases["teste_id"] == teste_id) & (fases["parceiro"] == parceiro)
        fases_teste = fases[filtro]
        fase_principal = int(decisao["fase_principal"])
        resumo_teste = resumo[
            (resumo["teste_id"] == teste_id)
            & (resumo["parceiro"] == parceiro)
            & (resumo["fase_id"] == fase_principal)
        ]
        dados_teste = dados[
            (dados["teste_id"] == teste_id) & (dados["parceiro"] == parceiro)
        ].sort_values(["data", "grupo"])
        conteudo_hash = dados_teste.to_csv(index=False).encode("utf-8")
        recomendada = str(decisao["variante_recomendada"])
        linha_recomendada = resumo_teste[resumo_teste["grupo"] == recomendada]
        resultado_direto = (
            float(linha_recomendada["margem"].iloc[0])
            if not linha_recomendada.empty
            else None
        )
        controle = "Grupo 1"
        variantes = ", ".join(
            resumo_teste.loc[resumo_teste["grupo"] != controle, "grupo"].astype(str)
        ) or "Sem variante"
        registros.append(
            {
                "id_analise": f"AB-{indice + 1:03d}",
                "data_analise": date.today(),
                "teste_id": teste_id,
                "nome_teste": f"Teste A/B de cashback - {parceiro}",
                "descricao": decisao["descricao"],
                "parceiro": parceiro,
                "periodo": (
                    f"{fases_teste['data_inicio'].min():%d/%m/%Y} a "
                    f"{fases_teste['data_fim'].max():%d/%m/%Y}"
                ),
                "controle": controle,
                "variantes": variantes,
                "metrica_principal": "Resultado direto",
                "variante_recomendada": recomendada,
                "resultado_direto": resultado_direto,
                "decisao": decisao["decisao"],
                "status_qualidade": _status_qualidade(decisao),
                "alertas": decisao["alertas"],
                "caminho_relatorio": f"{teste_id}__{_identificador_texto(parceiro)}/relatorio.md",
                "hash_dados": hashlib.sha256(conteudo_hash).hexdigest()[:16],
                "versao": "1.0",
            }
        )
    return pd.DataFrame(registros)


def _identificador_texto(valor: str) -> str:
    """Normaliza texto para compor caminhos sem caracteres especiais."""

    normalizado = unicodedata.normalize("NFKD", valor)
    sem_acentos = "".join(
        caractere for caractere in normalizado if not unicodedata.combining(caractere)
    )
    texto = re.sub(r"[^a-z0-9]+", "_", sem_acentos.lower()).strip("_")
    return texto or "teste"


def _criar_dashboard(
    pasta: Workbook,
    decisoes: pd.DataFrame,
    fases: pd.DataFrame,
    resumo: pd.DataFrame,
    tabelas_usadas: set[str],
) -> None:
    """Cria a visao executiva da pasta de trabalho."""

    aba = pasta.create_sheet("Dashboard")
    _titulo(
        aba,
        "Analise de Testes A/B de Cashback | Meliuz",
        "Metricas calculadas pelo Python; conclusoes narrativas inseridas pelo agente.",
        11,
    )

    principal = _resumo_fases_principais(resumo, decisoes)
    recomendadas = principal.merge(
        decisoes[["teste_id", "parceiro", "variante_recomendada", "status"]],
        on=["teste_id", "parceiro"],
        how="left",
    )
    recomendadas = recomendadas[
        (recomendadas["grupo"] == recomendadas["variante_recomendada"])
        & (recomendadas["status"] == "escalar")
    ]
    indicadores = [
        ("Testes analisados", len(decisoes), COR_VERDE_CLARO),
        (
            "Variantes avaliadas",
            int(len(principal)),
            COR_VERDE_CLARO,
        ),
        (
            "Resultado direto recomendado",
            float(recomendadas["margem"].sum()),
            COR_VERDE_CLARO,
        ),
        (
            "Alertas de qualidade",
            int((decisoes.apply(_status_qualidade, axis=1) != "Dados consistentes").sum()),
            COR_AMARELO_CLARO,
        ),
    ]
    for indice, (rotulo, valor, cor) in enumerate(indicadores):
        coluna = 1 + indice * 3
        aba.merge_cells(start_row=5, start_column=coluna, end_row=5, end_column=coluna + 1)
        aba.merge_cells(start_row=6, start_column=coluna, end_row=7, end_column=coluna + 1)
        aba.cell(5, coluna, rotulo)
        aba.cell(5, coluna).font = Font(bold=True, color=COR_CINZA)
        aba.cell(6, coluna, valor)
        aba.cell(6, coluna).font = Font(bold=True, size=20, color=COR_TEXTO)
        if rotulo == "Resultado direto recomendado":
            aba.cell(6, coluna).number_format = '"R$" #,##0.00'
        for linha in range(5, 8):
            for coluna_cartao in range(coluna, coluna + 2):
                aba.cell(linha, coluna_cartao).fill = PatternFill("solid", fgColor=cor)
                aba.cell(linha, coluna_cartao).alignment = Alignment(vertical="center")

    tabela_decisoes = _dados_decisoes(decisoes, fases, resumo)
    _secao(aba, 9, "Decisoes executivas", len(tabela_decisoes.columns))
    fim_decisoes, _ = _escrever_tabela(
        aba,
        tabela_decisoes,
        10,
        1,
        "DecisoesExecutivas",
        tabelas_usadas,
    )
    for linha in range(11, fim_decisoes + 1):
        aba.row_dimensions[linha].height = 42
        cor_status = (
            COR_VERDE_CLARO
            if aba.cell(linha, 8).value == "escalar"
            else COR_AMARELO_CLARO
        )
        cor_qualidade = (
            COR_VERDE_CLARO
            if aba.cell(linha, 10).value == "Dados consistentes"
            else COR_AMARELO_CLARO
        )
        aba.cell(linha, 8).fill = PatternFill("solid", fgColor=cor_status)
        aba.cell(linha, 10).fill = PatternFill("solid", fgColor=cor_qualidade)

    base_graficos = principal[
        ["configuracao", "margem_dia", "vendas_dia", "taxa_cashback", "taxa_margem"]
    ]
    linha_graficos = fim_decisoes + 3
    _secao(aba, linha_graficos, "Base dos graficos da fase principal", len(base_graficos.columns))
    inicio_tabela = linha_graficos + 1
    fim_graficos, _ = _escrever_tabela(
        aba,
        base_graficos,
        inicio_tabela,
        1,
        "BaseGraficosDashboard",
        tabelas_usadas,
    )
    _grafico_barras(
        aba,
        inicio_tabela,
        fim_graficos,
        1,
        2,
        "Resultado direto diario da fase principal",
        "N5",
    )
    _grafico_barras(
        aba,
        inicio_tabela,
        fim_graficos,
        1,
        3,
        "GMV diario da fase principal",
        "N22",
    )

    linha_conclusoes = fim_graficos + 3
    _secao(aba, linha_conclusoes, "Sintese do agente", 11)
    linha_cabecalho = linha_conclusoes + 1
    aba.merge_cells(
        start_row=linha_cabecalho,
        start_column=3,
        end_row=linha_cabecalho,
        end_column=11,
    )
    for coluna, texto in [(1, "Teste"), (2, "Parceiro"), (3, "Conclusao do agente")]:
        aba.cell(linha_cabecalho, coluna, texto)
        aba.cell(linha_cabecalho, coluna).font = Font(bold=True, color=COR_BRANCO)
        aba.cell(linha_cabecalho, coluna).fill = PatternFill("solid", fgColor=COR_VERDE)
    for coluna in range(3, 12):
        aba.cell(linha_cabecalho, coluna).fill = PatternFill("solid", fgColor=COR_VERDE)

    for deslocamento, decisao in enumerate(decisoes.itertuples(), start=1):
        linha = linha_cabecalho + deslocamento
        aba.cell(linha, 1, decisao.teste_id)
        aba.cell(linha, 2, decisao.parceiro)
        aba.merge_cells(start_row=linha, start_column=3, end_row=linha, end_column=11)
        aba.cell(linha, 3, "Aguardando revisao do agente.")
        cor = COR_VERDE_CLARO if deslocamento % 2 else COR_BRANCO
        for coluna in range(1, 12):
            aba.cell(linha, coluna).fill = PatternFill("solid", fgColor=cor)
            aba.cell(linha, coluna).alignment = Alignment(wrap_text=True, vertical="top")
        aba.row_dimensions[linha].height = 68
    fim_conclusoes = linha_cabecalho + len(decisoes)

    linha_criterios = fim_conclusoes + 3
    _secao(aba, linha_criterios, "Criterios de decisao", 11)
    criterios = [
        "Escalar somente quando o intervalo de 95% do resultado direto sustenta superioridade.",
        "Reexecutar testes com mudancas persistentes de cashback ou comissao.",
        "Usar compradores e GMV como guardrails; sem exposicao, nao sao taxas de conversao.",
    ]
    for deslocamento, criterio in enumerate(criterios, start=1):
        aba.merge_cells(
            start_row=linha_criterios + deslocamento,
            start_column=1,
            end_row=linha_criterios + deslocamento,
            end_column=11,
        )
        aba.cell(linha_criterios + deslocamento, 1, criterio)
        aba.cell(linha_criterios + deslocamento, 1).alignment = Alignment(wrap_text=True)
    aba.freeze_panes = "A10"
    _ajustar_colunas(aba, limite=38)


def _criar_aba_teste(
    pasta: Workbook,
    decisao: pd.Series,
    fases: pd.DataFrame,
    resumo: pd.DataFrame,
    comparacoes: pd.DataFrame,
    economia: pd.DataFrame,
    nomes_abas: set[str],
    tabelas_usadas: set[str],
) -> None:
    """Cria uma pagina executiva para um teste e parceiro."""

    teste_id = str(decisao["teste_id"])
    parceiro = str(decisao["parceiro"])
    filtro = (fases["teste_id"] == teste_id) & (fases["parceiro"] == parceiro)
    fases_teste = fases[filtro].copy()
    resumo_teste = resumo[
        (resumo["teste_id"] == teste_id) & (resumo["parceiro"] == parceiro)
    ].copy()
    comparacoes_teste = comparacoes[
        (comparacoes["teste_id"] == teste_id) & (comparacoes["parceiro"] == parceiro)
    ].copy()
    nome = _nome_aba(parceiro, nomes_abas)
    aba = pasta.create_sheet(nome)
    periodo = (
        f"Teste: {teste_id} | Periodo analisado: "
        f"{fases_teste['data_inicio'].min():%d/%m/%Y} a "
        f"{fases_teste['data_fim'].max():%d/%m/%Y}"
    )
    _titulo(aba, f"{parceiro} | Resultado do teste A/B", periodo, 13)

    cor_status = COR_VERDE_CLARO if decisao["status"] == "escalar" else COR_AMARELO_CLARO
    aba.merge_cells("A5:D5")
    aba["A5"] = "RECOMENDACAO"
    aba["A5"].font = Font(bold=True, color=COR_CINZA)
    aba.merge_cells("A6:D8")
    aba["A6"] = str(decisao["decisao"])
    aba["A6"].font = Font(bold=True, size=11, color=COR_TEXTO)
    aba["A6"].alignment = Alignment(vertical="center", wrap_text=True)
    for linha in range(5, 9):
        for coluna in range(1, 5):
            aba.cell(linha, coluna).fill = PatternFill("solid", fgColor=cor_status)

    aba.merge_cells("E5:H5")
    aba["E5"] = "QUALIDADE DO TESTE"
    aba["E5"].font = Font(bold=True, color=COR_CINZA)
    aba.merge_cells("E6:H8")
    aba["E6"] = _status_qualidade(decisao)
    aba["E6"].font = Font(bold=True, size=11, color=COR_TEXTO)
    aba["E6"].alignment = Alignment(vertical="center", wrap_text=True)
    cor_qualidade = (
        COR_VERDE_CLARO
        if _status_qualidade(decisao) == "Dados consistentes"
        else COR_AMARELO_CLARO
    )
    for linha in range(5, 9):
        for coluna in range(5, 9):
            aba.cell(linha, coluna).fill = PatternFill("solid", fgColor=cor_qualidade)

    aba.merge_cells("I5:M5")
    aba["I5"] = "JUSTIFICATIVA PRINCIPAL"
    aba["I5"].font = Font(bold=True, color=COR_CINZA)
    aba.merge_cells("I6:M8")
    aba["I6"] = "Aguardando revisao do agente."
    aba["I6"].alignment = Alignment(vertical="center", wrap_text=True)
    for linha in range(5, 9):
        for coluna in range(9, 14):
            aba.cell(linha, coluna).fill = PatternFill("solid", fgColor=COR_CINZA_CLARO)

    resumo_exibicao = _resumo_executivo_teste(
        resumo_teste,
        comparacoes_teste,
        decisao,
    )
    _secao(aba, 10, "Resumo da fase principal", len(resumo_exibicao.columns))
    inicio_resumo = 11
    fim_resumo, _ = _escrever_tabela(
        aba,
        resumo_exibicao,
        inicio_resumo,
        1,
        f"Resumo_{teste_id}_{parceiro}",
        tabelas_usadas,
    )
    _grafico_barras(
        aba,
        inicio_resumo,
        fim_resumo,
        1,
        7,
        "Resultado direto por grupo",
        "O5",
    )
    _grafico_barras(
        aba,
        inicio_resumo,
        fim_resumo,
        1,
        4,
        "GMV por grupo",
        "O22",
    )

    fase_principal = int(decisao["fase_principal"])
    comparacoes_exibicao = _comparacoes_executivas(
        comparacoes_teste[comparacoes_teste["fase_id"] == fase_principal]
    )
    linha_comparacoes = fim_resumo + 3
    _secao(
        aba,
        linha_comparacoes,
        "Evidencia estatistica da fase principal",
        max(1, len(comparacoes_exibicao.columns)),
    )
    fim_comparacoes, _ = _escrever_tabela(
        aba,
        comparacoes_exibicao,
        linha_comparacoes + 1,
        1,
        f"Comparacoes_{teste_id}_{parceiro}",
        tabelas_usadas,
    )

    linha_fases = fim_comparacoes + 3
    fases_exibicao = fases_teste[
        [
            "fase_id",
            "data_inicio",
            "data_fim",
            "dias",
            "quantidade_grupos",
            "configuracao_cashback",
            "taxa_comissao",
            "status_fase",
        ]
    ]
    _secao(aba, linha_fases, "Fases detectadas", len(fases_exibicao.columns))
    fim_fases, _ = _escrever_tabela(
        aba,
        fases_exibicao,
        linha_fases + 1,
        1,
        f"Fases_{teste_id}_{parceiro}",
        tabelas_usadas,
    )

    linha_leitura = fim_fases + 3
    _secao(aba, linha_leitura, "LEITURA GERENCIAL E RISCOS", 13)
    aba.merge_cells(
        start_row=linha_leitura + 1,
        start_column=1,
        end_row=linha_leitura + 4,
        end_column=13,
    )
    aba.cell(linha_leitura + 1, 1, "Aguardando revisao do agente.")
    aba.cell(linha_leitura + 1, 1).alignment = Alignment(wrap_text=True, vertical="top")

    linha_proximos = linha_leitura + 6
    _secao(aba, linha_proximos, "PROXIMOS PASSOS", 13)
    aba.merge_cells(
        start_row=linha_proximos + 1,
        start_column=1,
        end_row=linha_proximos + 4,
        end_column=13,
    )
    aba.cell(linha_proximos + 1, 1, "Aguardando revisao do agente.")
    aba.cell(linha_proximos + 1, 1).alignment = Alignment(wrap_text=True, vertical="top")
    aba.freeze_panes = "A11"
    _ajustar_colunas(aba)


def _criar_aba_dados(
    pasta: Workbook,
    nome: str,
    dados: pd.DataFrame,
    tabelas_usadas: set[str],
    colunas: list[str] | None = None,
) -> None:
    """Cria uma aba tabular para auditoria dos resultados."""

    aba = pasta.create_sheet(nome)
    tabela = dados[colunas].copy() if colunas else dados.copy()
    _titulo(aba, nome, "Dados calculados pelo pipeline Python.", max(8, len(tabela.columns)))
    _escrever_tabela(aba, tabela, 5, 1, f"Tabela_{nome}", tabelas_usadas)
    aba.freeze_panes = "A6"
    aba.auto_filter.ref = aba.dimensions
    _ajustar_colunas(aba)


def _criar_metodologia(pasta: Workbook) -> None:
    """Registra as definicoes usadas para interpretar a planilha."""

    aba = pasta.create_sheet("Metodologia")
    _titulo(
        aba,
        "Metodologia",
        "Definicoes e criterios usados pela analise automatizada.",
    )
    metricas = [
        ("Resultado direto", "Comissao - cashback; metrica principal para decisao."),
        ("Taxa de margem", "Resultado direto / GMV; compara eficiencia financeira."),
        ("Compradores", "Volume observado; guardrail sem interpretacao de conversao."),
        ("GMV", "Valor total vendido; guardrail de atividade comercial."),
        ("Ticket medio", "GMV / compradores; indica mudancas no valor por compra."),
    ]
    regras = [
        (
            "Guardrails",
            "Compradores e GMV sao comparados ao controle, mas nao representam "
            "conversao sem dados de exposicao.",
        ),
        (
            "Fases",
            "Mudancas persistentes de cashback ou comissao por pelo menos tres "
            "dias criam uma nova fase.",
        ),
        (
            "Inferencia",
            "Bootstrap e permutacao usam diferencas diarias pareadas dentro de "
            "cada fase.",
        ),
        (
            "Escala",
            "Uma variante so e escalada quando o intervalo de 95% da diferenca "
            "de resultado direto sustenta a superioridade.",
        ),
        (
            "Reexecucao",
            "Testes com mais de uma fase devem ser reexecutados com tratamentos "
            "fixos.",
        ),
        (
            "Projecao",
            "A projecao de 30 dias pressupoe distribuicao equilibrada de trafego "
            "entre os grupos.",
        ),
    ]
    limitacoes = [
        ("Exposicao", "Nao ha numero de usuarios expostos por grupo nem taxa de conversao."),
        ("Agregacao", "Os registros diarios nao permitem analisar comportamento individual."),
        ("Custos", "Resultado direto nao inclui impostos, cancelamentos ou outros custos."),
        ("Projecao", "A escala para 100% pressupoe distribuicao equilibrada de trafego."),
        ("Causalidade", "A leitura causal depende da correta randomizacao fora do dataset."),
    ]

    linha = 5
    for titulo, conteudo in [
        ("Definicoes das metricas", metricas),
        ("Regras de decisao", regras),
        ("Limitacoes do dataset", limitacoes),
    ]:
        _secao(aba, linha, titulo, 8)
        linha += 1
        for tema, explicacao in conteudo:
            indice = linha
            linha += 1
            aba.cell(indice, 1, tema)
            aba.cell(indice, 1).font = Font(bold=True, color=COR_TEXTO)
            aba.merge_cells(start_row=indice, start_column=2, end_row=indice, end_column=8)
            aba.cell(indice, 2, explicacao)
            aba.cell(indice, 2).alignment = Alignment(wrap_text=True, vertical="top")
            aba.row_dimensions[indice].height = 34
        linha += 1

    _secao(aba, linha, "Responsabilidades", 8)
    responsabilidades = [
        ("Python", "Calcula metricas, fases, intervalos, alertas e decisao deterministica."),
        ("Agente", "Interpreta os resultados e redige conclusoes sem alterar os numeros."),
        ("Fonte", "Arquivos CSV encontrados na entrada informada ao pipeline."),
    ]
    for tema, explicacao in responsabilidades:
        linha += 1
        aba.cell(linha, 1, tema)
        aba.cell(linha, 1).font = Font(bold=True, color=COR_TEXTO)
        aba.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=8)
        aba.cell(linha, 2, explicacao)
        aba.cell(linha, 2).alignment = Alignment(wrap_text=True, vertical="top")
        aba.row_dimensions[linha].height = 34
    _ajustar_colunas(aba)
    aba.column_dimensions["A"].width = 22
    aba.column_dimensions["B"].width = 70


def gerar_planilha_excel(
    destino: str | Path,
    dados: pd.DataFrame,
    fases: pd.DataFrame,
    resumo: pd.DataFrame,
    comparacoes: pd.DataFrame,
    economia: pd.DataFrame,
    decisoes: pd.DataFrame,
    acompanhamento: pd.DataFrame,
    alertas_qualidade: pd.DataFrame,
) -> Path:
    """Recria a planilha completa e substitui o arquivo anterior com seguranca."""

    caminho = Path(destino)
    caminho.parent.mkdir(parents=True, exist_ok=True)
    temporario = caminho.with_name(f".{caminho.stem}.temporario.xlsx")

    pasta = Workbook()
    pasta.remove(pasta.active)
    nomes_abas: set[str] = set()
    tabelas_usadas: set[str] = set()

    _criar_dashboard(pasta, decisoes, fases, resumo, tabelas_usadas)
    nomes_abas.add("dashboard")
    acompanhamento_detalhado = _acompanhamento_executivo(
        decisoes,
        fases,
        resumo,
        dados,
    )
    _criar_aba_dados(
        pasta,
        "Acompanhamento",
        acompanhamento_detalhado,
        tabelas_usadas,
    )
    _criar_aba_dados(pasta, "Consolidado", resumo, tabelas_usadas)
    _criar_aba_dados(pasta, "Fases", fases, tabelas_usadas)
    _criar_aba_dados(pasta, "Comparacoes", comparacoes, tabelas_usadas)
    _criar_aba_dados(pasta, "Economia", economia, tabelas_usadas)
    _criar_aba_dados(pasta, "Qualidade", alertas_qualidade, tabelas_usadas)
    _criar_aba_dados(pasta, "Dados", dados, tabelas_usadas)
    _criar_metodologia(pasta)
    nomes_abas.update(aba.title.lower() for aba in pasta.worksheets)

    for _, decisao in decisoes.iterrows():
        _criar_aba_teste(
            pasta,
            decisao,
            fases,
            resumo,
            comparacoes,
            economia,
            nomes_abas,
            tabelas_usadas,
        )

    try:
        pasta.save(temporario)
        temporario.replace(caminho)
    finally:
        if temporario.exists():
            temporario.unlink()
    return caminho


def _normalizar_titulo(texto: str) -> str:
    """Remove acentos e variacoes de caixa para localizar secoes Markdown."""

    normalizado = unicodedata.normalize("NFKD", texto)
    return "".join(
        caractere for caractere in normalizado if not unicodedata.combining(caractere)
    ).lower().strip()


def _extrair_secoes_markdown(conteudo: str) -> dict[str, str]:
    """Extrai secoes narrativas de um relatorio revisado pelo agente."""

    titulos = list(re.finditer(r"^#{2,3}\s+(.+?)\s*$", conteudo, flags=re.MULTILINE))
    secoes: dict[str, str] = {}
    for indice, titulo in enumerate(titulos):
        inicio = titulo.end()
        fim = titulos[indice + 1].start() if indice + 1 < len(titulos) else len(conteudo)
        texto = conteudo[inicio:fim].strip()
        texto = re.sub(r"\*\*|`", "", texto)
        texto = re.sub(r"\n{3,}", "\n\n", texto)
        secoes[_normalizar_titulo(titulo.group(1))] = texto
    return secoes


def _carregar_narrativas(diretorio_relatorios: Path) -> dict[tuple[str, str], dict[str, str]]:
    """Relaciona as narrativas revisadas aos respectivos testes."""

    narrativas: dict[tuple[str, str], dict[str, str]] = {}
    for caminho_contexto in diretorio_relatorios.rglob("contexto_llm.json"):
        caminho_relatorio = caminho_contexto.with_name("relatorio.md")
        if not caminho_relatorio.exists():
            continue
        contexto = json.loads(caminho_contexto.read_text(encoding="utf-8"))
        decisao = contexto.get("decisao_deterministica", {})
        teste_id = str(decisao.get("teste_id", ""))
        parceiro = str(decisao.get("parceiro", ""))
        if not teste_id or not parceiro:
            continue
        secoes = _extrair_secoes_markdown(
            caminho_relatorio.read_text(encoding="utf-8")
        )
        leitura = secoes.get("leitura executiva", "")
        proximos = next(
            (
                texto
                for titulo, texto in secoes.items()
                if "proximos passos" in titulo
            ),
            "",
        )
        if leitura or proximos:
            narrativas[(teste_id, parceiro)] = {
                "leitura": leitura or "Aguardando revisao do agente.",
                "proximos": proximos or "Aguardando revisao do agente.",
            }
    return narrativas


def atualizar_planilha_com_narrativas(
    planilha: str | Path,
    diretorio_relatorios: str | Path,
) -> int:
    """Insere textos do agente na planilha sem modificar as metricas calculadas."""

    caminho = Path(planilha)
    narrativas = _carregar_narrativas(Path(diretorio_relatorios))
    if not narrativas:
        return 0

    pasta = load_workbook(caminho)
    atualizados = 0
    for aba in pasta.worksheets:
        subtitulo = str(aba["A3"].value or "")
        correspondencia = re.search(r"Teste:\s*([^|]+)", subtitulo)
        if not correspondencia:
            continue
        teste_id = correspondencia.group(1).strip()
        item = next(
            (
                narrativa
                for (identificador, _), narrativa in narrativas.items()
                if identificador == teste_id
            ),
            None,
        )
        if item is None:
            continue
        aba["I6"] = item["leitura"]
        aba["I6"].alignment = Alignment(wrap_text=True, vertical="center")
        for linha in range(1, aba.max_row + 1):
            rotulo = str(aba.cell(linha, 1).value or "")
            if rotulo == "LEITURA GERENCIAL E RISCOS":
                aba.cell(linha + 1, 1, item["leitura"])
            elif rotulo == "PROXIMOS PASSOS":
                aba.cell(linha + 1, 1, item["proximos"])
        atualizados += 1

    dashboard = pasta["Dashboard"]
    for linha in range(1, dashboard.max_row + 1):
        if dashboard.cell(linha, 3).value != "Conclusao do agente":
            continue
        for linha_dados in range(linha + 1, dashboard.max_row + 1):
            teste_id = str(dashboard.cell(linha_dados, 1).value or "")
            parceiro = str(dashboard.cell(linha_dados, 2).value or "")
            item = narrativas.get((teste_id, parceiro))
            if item is None:
                break
            dashboard.cell(linha_dados, 3, item["leitura"])
            dashboard.cell(linha_dados, 3).alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )
        break

    temporario = caminho.with_name(f".{caminho.stem}.narrativas.xlsx")
    try:
        pasta.save(temporario)
        temporario.replace(caminho)
    finally:
        if temporario.exists():
            temporario.unlink()
    return atualizados


def executar_atualizacao() -> None:
    """Disponibiliza a etapa narrativa como um comando simples para o agente."""

    analisador = argparse.ArgumentParser(
        description="Insere na planilha as conclusoes dos relatorios revisados."
    )
    analisador.add_argument(
        "--planilha",
        default="relatorios/Resultados_Testes_AB_Meliuz.xlsx",
    )
    analisador.add_argument("--relatorios", default="relatorios")
    argumentos = analisador.parse_args()
    quantidade = atualizar_planilha_com_narrativas(
        argumentos.planilha,
        argumentos.relatorios,
    )
    print(f"Narrativas atualizadas em {quantidade} aba(s) de teste.")


if __name__ == "__main__":
    executar_atualizacao()
